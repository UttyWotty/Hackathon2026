"""
Risk Tower Excel Sheet Generation
==================================

Creates the Risk Tower sheet in RunRate Excel reports with equipment risk rankings,
trend analysis, and RAG status visualization.
"""

import logging

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from ..core.risk_tower import calculate_risk_tower

logger = logging.getLogger(__name__)

# Style constants
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
RAG_RED_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
RAG_AMBER_FILL = PatternFill(
    start_color="FFB347", end_color="FFB347", fill_type="solid"
)
RAG_GREEN_FILL = PatternFill(
    start_color="77DD77", end_color="77DD77", fill_type="solid"
)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_risk_tower_sheet(
    ws: Worksheet,
    df_result: pd.DataFrame,
    weeks: int = 4,
) -> None:
    """
    Create Risk Tower sheet with equipment risk rankings.

    Args:
        ws: Worksheet to write to
        df_result: Processed session DataFrame with all metrics
        weeks: Number of weeks for rolling window analysis
    """
    logger.info("Creating Risk Tower sheet...")

    # Calculate Risk Tower data
    risk_tower_df = calculate_risk_tower(df_result, weeks=weeks)

    if risk_tower_df.empty:
        ws["A1"] = "No data available for Risk Tower analysis"
        logger.warning("No data for Risk Tower - empty DataFrame")
        return

    # Write title
    ws["A1"] = "RISK TOWER - Equipment Run Rate Risk Analysis"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")

    ws["A2"] = (
        f"Rolling {weeks}-Week Analysis | Sorted by Risk Score (Lowest = Highest Risk)"
    )
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A2:H2")

    # Define columns to display
    display_columns = [
        ("EQUIPMENT_CODE", "Equipment"),
        ("RAG_STATUS", "Status"),
        ("RISK_SCORE", "Risk Score"),
        ("STABILITY_INDEX", "Stability %"),
        ("PRIMARY_RISK_FACTOR", "Primary Risk Factor"),
        ("FIRST_WEEK_STABILITY", "First Week %"),
        ("LAST_WEEK_STABILITY", "Last Week %"),
        ("TREND_CHANGE", "Trend Δ"),
        ("MTTR", "MTTR (min)"),
        ("MTBF", "MTBF (min)"),
        ("STOP_EVENTS", "Stop Events"),
        ("WEEKS_ANALYZED", "Weeks"),
    ]

    # Write headers (row 4)
    header_row = 4
    for col_idx, (col_name, display_name) in enumerate(display_columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=display_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Write data rows
    data_start_row = header_row + 1
    for row_idx, row_data in enumerate(
        risk_tower_df.itertuples(index=False), start=data_start_row
    ):
        row_dict = risk_tower_df.iloc[row_idx - data_start_row].to_dict()

        for col_idx, (col_name, _) in enumerate(display_columns, start=1):
            value = row_dict.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Apply RAG coloring to Status column
            if col_name == "RAG_STATUS":
                if value == "Red":
                    cell.fill = RAG_RED_FILL
                elif value == "Amber":
                    cell.fill = RAG_AMBER_FILL
                elif value == "Green":
                    cell.fill = RAG_GREEN_FILL

            # Format numeric columns
            if col_name in [
                "STABILITY_INDEX",
                "FIRST_WEEK_STABILITY",
                "LAST_WEEK_STABILITY",
            ]:
                cell.number_format = "0.0"
            elif col_name in ["MTTR", "MTBF", "TREND_CHANGE"]:
                cell.number_format = "0.00"

    # Set column widths
    column_widths = {
        "A": 15,  # Equipment
        "B": 10,  # Status
        "C": 12,  # Risk Score
        "D": 12,  # Stability %
        "E": 20,  # Primary Risk Factor
        "F": 14,  # First Week %
        "G": 14,  # Last Week %
        "H": 10,  # Trend Δ
        "I": 12,  # MTTR
        "J": 12,  # MTBF
        "K": 12,  # Stop Events
        "L": 10,  # Weeks
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Add legend
    legend_row = data_start_row + len(risk_tower_df) + 2
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)

    ws.cell(row=legend_row + 1, column=1, value="●").fill = RAG_GREEN_FILL
    ws.cell(row=legend_row + 1, column=2, value="Stable (≥70% Stability)")

    ws.cell(row=legend_row + 2, column=1, value="●").fill = RAG_AMBER_FILL
    ws.cell(row=legend_row + 2, column=2, value="Moderate (50-70% Stability)")

    ws.cell(row=legend_row + 3, column=1, value="●").fill = RAG_RED_FILL
    ws.cell(row=legend_row + 3, column=2, value="Critical (<50% Stability)")

    # Add risk factor definitions
    definitions_row = legend_row + 5
    ws.cell(row=definitions_row, column=1, value="Risk Factor Definitions:").font = (
        Font(bold=True)
    )

    risk_definitions = [
        ("Declining Trend", ">5% stability drop from first to last active week"),
        ("High MTTR", "Mean Time To Repair > 1.2× average (slow recovery)"),
        ("Frequent Stops", "Mean Time Between Failures < 0.8× average"),
        ("Critical Stability", "<50% of run time in normal production"),
        ("Moderate Stability", "50-70% of run time in normal production"),
        ("Stable", "≥70% of run time in normal production"),
    ]

    for idx, (factor, definition) in enumerate(risk_definitions, start=1):
        ws.cell(row=definitions_row + idx, column=1, value=factor).font = Font(
            bold=True
        )
        ws.cell(row=definitions_row + idx, column=2, value=definition)

    logger.info(f"Risk Tower sheet created with {len(risk_tower_df)} equipment entries")


def get_risk_summary_stats(df_result: pd.DataFrame, weeks: int = 4) -> dict:
    """
    Get summary statistics from Risk Tower analysis.

    Args:
        df_result: Processed session DataFrame
        weeks: Number of weeks for analysis

    Returns:
        Dictionary with summary stats
    """
    risk_tower_df = calculate_risk_tower(df_result, weeks=weeks)

    if risk_tower_df.empty:
        return {
            "total_equipment": 0,
            "red_count": 0,
            "amber_count": 0,
            "green_count": 0,
            "declining_count": 0,
            "avg_stability": 0,
        }

    return {
        "total_equipment": len(risk_tower_df),
        "red_count": (risk_tower_df["RAG_STATUS"] == "Red").sum(),
        "amber_count": (risk_tower_df["RAG_STATUS"] == "Amber").sum(),
        "green_count": (risk_tower_df["RAG_STATUS"] == "Green").sum(),
        "declining_count": risk_tower_df["IS_DECLINING"].sum(),
        "avg_stability": round(risk_tower_df["STABILITY_INDEX"].mean(), 1),
    }
