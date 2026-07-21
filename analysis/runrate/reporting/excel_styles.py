"""
Excel styling utilities for RunRate reports.

Provides standardized styles, colors, and formatting for Excel reports.
"""

from openpyxl.styles import Border, Font, PatternFill, Side


class ExcelStyles:
    """
    Centralized Excel styling configuration for RunRate reports.

    Provides consistent colors, fonts, and borders across all Excel reports.
    """

    # Color fills
    YELLOW_FILL = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    GREY_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    DARK_BLUE_FILL = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    LIGHT_BLUE_FILL = PatternFill(
        start_color="E1F5FE", end_color="E1F5FE", fill_type="solid"
    )
    GREEN_FILL = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )

    # Fonts
    WHITE_FONT = Font(color="FFFFFF", bold=True)
    BLACK_FONT = Font(bold=True)
    REGULAR_FONT = Font(bold=False)

    # Borders
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @classmethod
    def get_header_style(cls):
        """Get style for header cells."""
        return {
            "fill": cls.DARK_BLUE_FILL,
            "font": cls.WHITE_FONT,
            "border": cls.THIN_BORDER,
        }

    @classmethod
    def get_data_style(cls):
        """Get style for data cells."""
        return {
            "border": cls.THIN_BORDER,
        }

    @classmethod
    def get_highlight_style(cls, color="yellow"):
        """Get highlighting style for special cells."""
        fill_map = {
            "yellow": cls.YELLOW_FILL,
            "red": cls.RED_FILL,
            "grey": cls.GREY_FILL,
            "green": cls.GREEN_FILL,
            "dark_blue": cls.DARK_BLUE_FILL,
        }
        font = cls.WHITE_FONT if color == "dark_blue" else cls.BLACK_FONT
        return {
            "fill": fill_map.get(color, cls.GREY_FILL),
            "font": font,
            "border": cls.THIN_BORDER,
        }


def apply_cell_style(cell, fill=None, font=None, border=None):
    """
    Apply styling to a single cell.

    Args:
        cell: OpenPyXL cell object
        fill: PatternFill object (optional)
        font: Font object (optional)
        border: Border object (optional)
    """
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if border:
        cell.border = border


def create_header_cell(worksheet, cell_ref, value, style_dict=None):
    """
    Create and style a header cell.

    Args:
        worksheet: OpenPyXL worksheet object
        cell_ref: Cell reference (e.g., "A1")
        value: Cell value
        style_dict: Optional dict with 'fill', 'font', 'border' keys

    Returns:
        The cell object
    """
    cell = worksheet[cell_ref]
    cell.value = value

    if style_dict is None:
        style_dict = ExcelStyles.get_header_style()

    apply_cell_style(
        cell,
        fill=style_dict.get("fill"),
        font=style_dict.get("font"),
        border=style_dict.get("border"),
    )

    return cell


def create_data_cell(worksheet, cell_ref, value, style_dict=None):
    """
    Create and style a data cell.

    Args:
        worksheet: OpenPyXL worksheet object
        cell_ref: Cell reference (e.g., "B1")
        value: Cell value
        style_dict: Optional dict with 'fill', 'font', 'border' keys

    Returns:
        The cell object
    """
    cell = worksheet[cell_ref]
    cell.value = value

    if style_dict is None:
        style_dict = ExcelStyles.get_data_style()

    apply_cell_style(
        cell,
        fill=style_dict.get("fill"),
        font=style_dict.get("font"),
        border=style_dict.get("border"),
    )

    return cell
