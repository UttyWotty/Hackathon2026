"""
Daily aggregation logic for RunRate trend analysis.

Aggregates session-level data into daily summaries for trend charts.
"""

from typing import Tuple

import pandas as pd


def aggregate_sessions_to_daily(df_result: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate session-level metrics to daily summaries.

    This function properly handles session-level metrics to avoid duplication:
    1. First aggregates by session (one record per session)
    2. Then aggregates sessions by day

    Args:
        df_result: DataFrame with shot-level data including session metrics

    Returns:
        DataFrame with daily aggregated metrics

    Raises:
        ValueError: If required columns are missing

    Example:
        >>> daily = aggregate_sessions_to_daily(df_sessions)
        >>> print(f"Created {len(daily)} daily records")
    """
    print("   📅 Aggregating data by day...")

    # Step 1: Get unique session-level metrics (one record per session)
    print("   📅 Step 1: Aggregating session-level metrics...")
    df_sessions = df_result.copy()

    # Calculate SHOT_DATE based on shot END time (LOCAL_SHOT_TIME + CT)
    # This ensures cross-day shots are assigned to the day they complete
    # For 999.9 shots, use start time (conservative approach)
    df_sessions["SHOT_END_TIME"] = pd.to_datetime(
        df_sessions["LOCAL_SHOT_TIME"]
    ) + pd.to_timedelta(
        df_sessions["ACTUAL_CT"].apply(lambda ct: 0 if ct >= 999.9 else ct), unit="s"
    )
    df_sessions["SHOT_DATE"] = df_sessions["SHOT_END_TIME"].dt.date

    # Get session-level aggregation first (one record per session)
    session_aggregation = (
        df_sessions.groupby(["EQUIPMENT_CODE", "SESSION_ID", "SHOT_DATE"])
        .agg(
            {
                "LOCAL_SHOT_TIME": ["min", "max", "count"],
                "TOTAL_STOPS": "first",  # Stop events per session (back-to-back = 1 event)
                "ACTUAL_CT": "mean",  # Average cycle time in session
                "SHOT_DIFF_SEC": lambda x: (
                    x[df_sessions.loc[x.index, "STOP"] == 1].mean()
                    if (df_sessions.loc[x.index, "STOP"] == 1).any()
                    else 0
                ),
                "PRODUCTION_TIME": "first",  # Take first value (same per session)
                "TOTAL_DOWN_TIME": "first",  # Take first value (same per session)
                "EFFICIENCY": "first",  # Take first value (same per session)
                "MTTR": "first",  # Take first value (same per session)
                "MTBF": "first",  # Take first value (same per session)
            }
        )
        .reset_index()
    )

    # Flatten column names for session aggregation
    session_aggregation.columns = [
        "EQUIPMENT_CODE",
        "SESSION_ID",
        "SHOT_DATE",
        "SESSION_START_TIME",
        "SESSION_END_TIME",
        "SESSION_SHOT_COUNT",
        "SESSION_TOTAL_STOPS",  # Individual stops per session
        "SESSION_AVG_CYCLE_TIME",
        "SESSION_AVG_STOP_DURATION_SEC",
        "SESSION_PRODUCTION_TIME",
        "SESSION_DOWNTIME",
        "SESSION_EFFICIENCY",
        "SESSION_MTTR",
        "SESSION_MTBF",
    ]

    print(f"   ✅ Session aggregation: {len(session_aggregation)} session records")

    # Step 2: Now aggregate sessions by day (correct aggregation without duplication)
    print("   📅 Step 2: Aggregating by day...")

    # Filter out very small sessions (likely cross-day artifacts) for MTTR/MTBF calculation
    # These single-shot sessions skew the daily averages
    main_sessions = session_aggregation[
        session_aggregation["SESSION_SHOT_COUNT"] >= 3
    ].copy()

    # If all sessions are filtered out (edge case), use all sessions
    if main_sessions.empty:
        main_sessions = session_aggregation.copy()

    daily_summary = (
        session_aggregation.groupby(["EQUIPMENT_CODE", "SHOT_DATE"])
        .agg(
            {
                "SESSION_START_TIME": "min",  # Earliest session start in day
                "SESSION_END_TIME": "max",  # Latest session end in day
                "SESSION_SHOT_COUNT": "sum",  # Total shots across all sessions in day
                "SESSION_TOTAL_STOPS": "sum",  # Total stops across all sessions in day
                "SESSION_AVG_CYCLE_TIME": "mean",  # Average cycle time across sessions
                "SESSION_AVG_STOP_DURATION_SEC": "mean",  # Average stop duration
                "SESSION_PRODUCTION_TIME": "sum",  # Sum session production times
                "SESSION_DOWNTIME": "sum",  # Sum session downtimes
                "SESSION_EFFICIENCY": "mean",  # Average efficiency across sessions
            }
        )
        .reset_index()
    )

    # Calculate MTTR/MTBF separately using only main sessions (excludes cross-day artifacts)
    mttr_mtbf_by_day = (
        main_sessions.groupby(["EQUIPMENT_CODE", "SHOT_DATE"])
        .agg(
            {
                "SESSION_MTTR": "mean",  # Average MTTR from main sessions only
                "SESSION_MTBF": "mean",  # Average MTBF from main sessions only
            }
        )
        .reset_index()
    )

    # Merge MTTR/MTBF back into daily summary
    daily_summary = daily_summary.merge(
        mttr_mtbf_by_day, on=["EQUIPMENT_CODE", "SHOT_DATE"], how="left"
    )

    # Rename columns for final output
    daily_summary.columns = [
        "EQUIPMENT_CODE",
        "SHOT_DATE",
        "START_TIME",
        "END_TIME",
        "TOTAL_SHOTS",
        "TOTAL_STOPS",
        "AVG_CYCLE_TIME",
        "AVG_STOP_DURATION_SEC",
        "TOTAL_PRODUCTION_TIME",
        "TOTAL_DOWNTIME",
        "AVG_EFFICIENCY",
        "MTTR",  # Already calculated correctly per session, now averaged
        "MTBF",  # Already calculated correctly per session, now averaged
    ]

    # Calculate time-based efficiency for Trends & Graphs sheet
    # Time efficiency = (production time / total time) * 100
    daily_summary["TIME_EFFICIENCY"] = (
        daily_summary["TOTAL_PRODUCTION_TIME"]
        / (daily_summary["TOTAL_PRODUCTION_TIME"] + daily_summary["TOTAL_DOWNTIME"])
        * 100
    ).round(1)

    # Replace AVG_EFFICIENCY with TIME_EFFICIENCY
    daily_summary["AVG_EFFICIENCY"] = daily_summary["TIME_EFFICIENCY"]

    # Add date string for display
    daily_summary["DATE_STR"] = daily_summary["SHOT_DATE"].astype(str)

    print(f"   ✅ Created daily summary with {len(daily_summary)} records")

    # Debug: Show sample stop counts to verify
    if not daily_summary.empty:
        sample_stops = daily_summary["TOTAL_STOPS"].head(3).tolist()
        print(f"   🔍 Debug - First 3 daily TOTAL_STOPS values: {sample_stops}")

    return daily_summary


def create_daily_data_table(
    worksheet,
    daily_summary: pd.DataFrame,
    start_row: int = 7,
) -> Tuple[int, int]:
    """
    Create daily data table with headers and data rows.

    Args:
        worksheet: OpenPyXL worksheet object
        daily_summary: DataFrame with daily aggregated data
        start_row: Starting row for the table (default 7)

    Returns:
        Tuple of (data_start_row, data_end_row)

    Example:
        >>> data_start, data_end = create_daily_data_table(ws, daily_df)
        >>> print(f"Data spans rows {data_start} to {data_end}")
    """
    from openpyxl.styles import Border, Font, PatternFill, Side

    # Define styles
    dark_blue_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    white_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        "Date",
        "Equipment",
        "Total Shots",
        "Total Stops",
        "MTTR (min)",
        "MTBF (min)",
        "Efficiency %",
        "Prod Time (min)",
        "Downtime (min)",
    ]

    for col, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=start_row, column=col)
        cell.value = header
        cell.fill = dark_blue_fill
        cell.font = white_font
        cell.border = thin_border

    # Data rows
    for row_idx, (_, row) in enumerate(daily_summary.iterrows(), start=start_row + 1):
        worksheet.cell(row=row_idx, column=1, value=row["DATE_STR"]).border = (
            thin_border
        )
        worksheet.cell(row=row_idx, column=2, value=row["EQUIPMENT_CODE"]).border = (
            thin_border
        )
        worksheet.cell(row=row_idx, column=3, value=int(row["TOTAL_SHOTS"])).border = (
            thin_border
        )
        worksheet.cell(row=row_idx, column=4, value=int(row["TOTAL_STOPS"])).border = (
            thin_border
        )
        worksheet.cell(row=row_idx, column=5, value=round(row["MTTR"], 2)).border = (
            thin_border
        )
        worksheet.cell(row=row_idx, column=6, value=round(row["MTBF"], 2)).border = (
            thin_border
        )
        worksheet.cell(
            row=row_idx, column=7, value=round(row["AVG_EFFICIENCY"], 1)
        ).border = thin_border
        worksheet.cell(
            row=row_idx, column=8, value=round(row["TOTAL_PRODUCTION_TIME"], 1)
        ).border = thin_border
        worksheet.cell(
            row=row_idx, column=9, value=round(row["TOTAL_DOWNTIME"], 1)
        ).border = thin_border

    data_start_row = start_row + 1
    data_end_row = start_row + len(daily_summary)

    return data_start_row, data_end_row
