"""
Excel report generation for RunRate analysis.

Creates comprehensive Excel reports with formulas, formatting, and charts.
This is the main entry point for report generation using modular components.
"""

from typing import List, Optional, Union

import pandas as pd
from openpyxl import Workbook

from .chart_builder import create_daily_trends_sheet
from .data_validation import validate_data_for_excel
from .excel_data_table import write_data_table_headers, write_data_table_rows
from .excel_formulas import update_summary_formulas, update_time_bucket_formulas
from .excel_header import (
    create_production_metrics_section,
    create_report_header,
    create_section_separator,
)
from .excel_risk_tower import create_risk_tower_sheet


def create_excel_report_with_formulas(
    df_result: pd.DataFrame,
    equipment_code: Optional[str] = None,
    date_range: Optional[Union[List[str], tuple]] = None,
    output_file: str = "production_report_with_formulas.xlsx",
) -> str:
    """
    Create comprehensive Excel report with formulas and formatting.

    This function generates a complete RunRate analysis report including:

    **Main Report Sheet:**
    - Equipment information header
    - Data units reference table
    - Detailed shot-by-shot data with formulas for:
        * Stop detection flags
        * Cumulative production time
        * Run duration calculations
        * Time buckets
    - Dynamic Excel formulas (not static values)

    **Trends & Graphs Sheet:**
    - Daily aggregated metrics
    - MTTR/MTBF trends
    - Efficiency trends over time
    - Production vs downtime charts

    **Risk Tower Sheet:**
    - Equipment risk rankings over 4-week rolling window
    - Stability Index, MTTR, MTBF per equipment
    - RAG status (Red/Amber/Green) visualization
    - Primary risk factor identification
    - Trend analysis (first vs last active week)

    **Summary Calculations:**
    - Total shots and sessions
    - Stop counts and efficiency
    - MTTR/MTBF metrics
    - Time to first downtime

    Args:
        df_result: Processed session DataFrame with all metrics
        equipment_code: Equipment identifier for report header
        date_range: [start_date, end_date] for report header
        output_file: Path for output Excel file

    Returns:
        str: Path to generated Excel file

    Raises:
        ValueError: If data validation fails
        IOError: If file cannot be written

    Example:
        >>> df_processed = process_sessions(raw_data)
        >>> output_path = create_excel_report_with_formulas(
        ...     df_processed,
        ...     equipment_code="EMA-4104",
        ...     date_range=["2024-01-01", "2024-12-31"],
        ...     output_file="runrate_report.xlsx"
        ... )
        >>> print(f"Report saved to: {output_path}")

    Notes:
        - Automatically validates and sanitizes data
        - Limits to 375K rows for Excel stability
        - Uses Excel formulas for dynamic calculations
        - Includes comprehensive formatting and styling

    Architecture:
        Uses modular components for maintainability:
        - data_validation: Data integrity and sanitization
        - excel_header: Report headers and structure
        - excel_data_table: Data writing and formatting
        - excel_formulas: Formula generation and updates
        - chart_builder: Trend charts and visualizations

    Refactored Implementation:
        This function now uses modular components extracted from the
        original implementation. Benefits include:
        ✓ Testable modules
        ✓ Clear separation of concerns
        ✓ Easier maintenance and updates
        ✓ Reusable components
    """
    print("📊 Starting Excel report generation...")

    # Step 1: Validate and sanitize data
    df_result = validate_data_for_excel(df_result)

    # Step 2: Create workbook and worksheets
    wb = Workbook()
    ws = wb.active
    if ws and hasattr(ws, "title") and ws.title == "Sheet":
        ws.title = "Production Report"

    # Create dedicated sheet for trends and graphs
    graphs_ws = wb.create_sheet("Trends & Graphs")
    print("📊 Created dedicated 'Trends & Graphs' worksheet")

    # Step 3: Create report header
    create_report_header(ws, equipment_code, date_range)

    # Step 4: Determine data table starting position
    main_data_table_start_row = 35  # After header sections
    shot_data_start_row = main_data_table_start_row + 1

    # Step 5: Create production metrics section (with placeholder formulas)
    shot_data_end_row = shot_data_start_row + min(375000, len(df_result)) - 1
    create_production_metrics_section(ws, shot_data_start_row, shot_data_end_row)

    # Step 6: Add section separator
    create_section_separator(ws, 70, "SUMMARY SECTION ENDS HERE")

    # Step 7: Write data table
    write_data_table_headers(ws, main_data_table_start_row)
    rows_written, actual_end_row = write_data_table_rows(
        ws, df_result, shot_data_start_row, max_rows=375000
    )

    print(f"📝 Data table spans rows {shot_data_start_row} to {actual_end_row}")

    # Step 8: Update formulas with actual data ranges
    update_summary_formulas(ws, shot_data_start_row, actual_end_row, df_result)

    # Step 9: Update time bucket analysis
    update_time_bucket_formulas(ws, 16, df_result)  # Time buckets start at row 16

    # Step 10: Create trends and graphs sheet
    create_daily_trends_sheet(graphs_ws, df_result, ws)

    # Step 11: Create Risk Tower sheet (4-week rolling analysis)
    risk_tower_ws = wb.create_sheet("Risk Tower")
    create_risk_tower_sheet(risk_tower_ws, df_result, weeks=4)
    print("📊 Created 'Risk Tower' worksheet (4-week analysis)")

    # Step 12: Save workbook
    try:
        wb.save(output_file)
        print(f"✅ Excel report saved successfully: {output_file}")
        return output_file
    except Exception as e:
        print(f"❌ Error saving Excel file: {e}")
        raise IOError(f"Failed to save Excel file: {e}")
